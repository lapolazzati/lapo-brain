"""
CEG Italia - Project Management Dashboard Builder
Builds a multi-sheet xlsx with:
  - Dashboard (summary + status)
  - Gantt (weekly bars across workstreams)
  - Workstreams (detail)
  - Action Items (granular tasks)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from datetime import date, timedelta

# ----- Config -----
TODAY = date(2026, 5, 7)
WEEK_START = TODAY - timedelta(days=TODAY.weekday())  # Monday May 4 2026
N_WEEKS = 20  # ~5 months
WEEKS = [WEEK_START + timedelta(weeks=i) for i in range(N_WEEKS)]

# Colors
COL_HEADER_BG = "1F4E78"
COL_HEADER_FG = "FFFFFF"
COL_SUBHEADER_BG = "D9E1F2"
COL_BAR_PRIORITY = "C00000"     # P1 critical
COL_BAR_HIGH = "ED7D31"          # P2
COL_BAR_MEDIUM = "FFC000"        # P3
COL_BAR_NORMAL = "548235"        # P4-P5
COL_BAR_FUTURE = "808080"        # not started yet / dependent
COL_TODAY_LINE = "FF0000"
COL_DONE_BG = "C6EFCE"
COL_BLOCKED_BG = "FFC7CE"
COL_INPROG_BG = "FFEB9C"

THIN = Side(border_style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT_NAME = "Arial"

# ----- Workstream data -----
# Each workstream: id, name, priority(P1-P6), owner, status, start_week_idx, end_week_idx, notes
WORKSTREAMS = [
    {
        "id": "WS1",
        "name": "Commercial Terms (Retainer + 7% Fee)",
        "priority": "P1",
        "owner": "Steve + Lapo",
        "status": "In Progress",
        "start_w": 0, "end_w": 2,
        "deadline": "2026-05-22",
        "notes": "BLOCKING: must close before any client-facing move. Term sheet covers retainer, fee basis, exclusivity scope (Italy vs Mediterranean), expense policy.",
    },
    {
        "id": "WS2",
        "name": "Voluntary Cleanup - Industrial Pipeline",
        "priority": "P2",
        "owner": "Lapo",
        "status": "In Progress",
        "start_w": 0, "end_w": 12,
        "deadline": "Ongoing",
        "notes": "Entry segment. Lawyer-led intros (Brenelli) + commercialisti channel. Focus: cement, chemical, fuel stations.",
    },
    {
        "id": "WS3",
        "name": "Cementificio Luca - First Live Prospect",
        "priority": "P2",
        "owner": "Lapo",
        "status": "Not Started",
        "start_w": 0, "end_w": 1,
        "deadline": "2026-05-14",
        "notes": "Confirm company name, contaminant type, site size. Could be first paid feasibility.",
    },
    {
        "id": "WS4",
        "name": "Brenelli - Voluntary Referral Pipeline",
        "priority": "P2",
        "owner": "Lapo",
        "status": "Not Started",
        "start_w": 0, "end_w": 2,
        "deadline": "2026-05-21",
        "notes": "Scope intent + collaboration structure with criminal env lawyer. Highest-leverage channel.",
    },
    {
        "id": "WS5",
        "name": "Mandated Cleanup - ARPA Approval Pathway",
        "priority": "P3",
        "owner": "Steve + Lapo",
        "status": "Not Started",
        "start_w": 1, "end_w": 8,
        "deadline": "2026-06-04",
        "notes": "Decide path: senior geologist vs. Luca institutional contact. Map CEG US regulatory dossier for ARPA adaptation. Don't pitch this segment until path is clear.",
    },
    {
        "id": "WS6",
        "name": "Entity Setup + Liability Insurance",
        "priority": "P3",
        "owner": "Lapo",
        "status": "Not Started",
        "start_w": 1, "end_w": 6,
        "deadline": "2026-06-07",
        "notes": "S.r.l. entity, RC professionale quote, commercialista consult, ~€100K capital plan. Required before Istanbul.",
    },
    {
        "id": "WS7",
        "name": "Satellite Intelligence Partnership",
        "priority": "P4",
        "owner": "Lapo",
        "status": "Not Started",
        "start_w": 2, "end_w": 8,
        "deadline": "2026-06-30",
        "notes": "Scoping call → coverage, commercial model. Before/after CEG project validation. Brenelli on legal framing.",
    },
    {
        "id": "WS8",
        "name": "RDC Local Manufacturing",
        "priority": "P4",
        "owner": "Lapo + CEG",
        "status": "Initial Contact Made",
        "start_w": 3, "end_w": 14,
        "deadline": "2026-08-15",
        "notes": "BioBlend + ISCO formulation specs, IP/licensing constraints, quality requirements.",
    },
    {
        "id": "WS9",
        "name": "Turkey / Istanbul Expansion",
        "priority": "P5",
        "owner": "Lapo",
        "status": "On Hold",
        "start_w": 4, "end_w": 16,
        "deadline": "2026-09-30",
        "notes": "DEPENDS ON WS1 + WS6. Don't initiate before terms + entity. Frame as regional partner, not CEG-branded.",
    },
]

# ----- Action items (granular tasks) -----
ACTIONS = [
    # (workstream_id, task, owner, due_date, status, notes)
    ("WS1", "Draft term sheet covering retainer (€4K/mo) + 7% introducer fee + Italy exclusivity + expense policy", "Lapo", "2026-05-15", "Not Started", "Define fee basis: CEG contract value vs Lapo-introduced portion only"),
    ("WS1", "Steve review + sign side letter", "Steve", "2026-05-22", "Not Started", "Blocking gate for all downstream activity"),

    ("WS2", "Build target list: cement + chemical + fuel station polluters (channel partner shortlist)", "Lapo", "2026-05-31", "In Progress", "Reference Channel Partners shortlist sheet"),
    ("WS2", "Define CEG voluntary-client offering (bench test entry, scoping call, prelim estimate)", "CEG", "2026-06-15", "Not Started", "Need pricing tiers and SLA for first contact"),
    ("WS2", "Set up commercialisti channel funnel (Di Salvo, Rizzi)", "Lapo", "2026-06-30", "Not Started", "Tax DD as early warning system for contaminated sites"),

    ("WS3", "Get full company name + contaminant type + approx site size from Luca", "Lapo", "2026-05-14", "Not Started", "Priority 3 in original doc"),
    ("WS3", "Send 1-pager + CEG capability dossier to Luca", "Lapo", "2026-05-21", "Not Started", "Adapt from CEG US materials"),
    ("WS3", "Schedule technical scoping call (CEG + Luca + Lapo)", "Lapo", "2026-05-28", "Not Started", "Confirm bench test feasibility"),

    ("WS4", "Initial scoping call with Brenelli", "Lapo", "2026-05-21", "Not Started", "Volume of voluntary remediation inquiries, willingness to refer"),
    ("WS4", "Propose referral / collaboration structure to Brenelli", "Lapo", "2026-06-04", "Not Started", "Fee share or co-branded approach"),

    ("WS5", "Identify senior geologist with active ARPA relationships", "Lapo", "2026-05-21", "Not Started", "Alternative path to Luca's institutional contact"),
    ("WS5", "Decision: geologist path vs. Luca institutional contact", "Steve + Lapo", "2026-05-21", "Not Started", "Priority 4 in original doc"),
    ("WS5", "Map CEG US regulatory dossiers usable for ARPA submission", "CEG", "2026-06-15", "Not Started", "Piano di Caratterizzazione, Analisi di Rischio, Piano di Bonifica equivalents"),
    ("WS5", "Identify Italian env-law specialist for documentation adaptation", "Lapo + Steve", "2026-06-30", "Not Started", "Different from Brenelli (criminal side)"),

    ("WS6", "Confirm whether CEG insurance extends to Italian projects", "Steve", "2026-05-21", "Not Started", "Critical input for our liability gap"),
    ("WS6", "Get RC professionale quote for env tech intermediary", "Lapo", "2026-05-31", "Not Started", "Italian broker; not executing party, only intermediary"),
    ("WS6", "Commercialista consult: S.r.l. vs holding structure (Italy + Turkey)", "Lapo", "2026-05-31", "Not Started", "Liability ring-fencing + intl expansion"),
    ("WS6", "Investment breakdown + funding allocation (~€100K)", "Lapo + Steve", "2026-06-07", "Not Started", "Who funds what"),
    ("WS6", "Open S.r.l. entity", "Lapo", "2026-06-30", "Not Started", "Triggered after structure decision"),

    ("WS7", "Scoping call with satellite contact (coverage, commercial model)", "Lapo", "2026-05-31", "Not Started", "Polluter ID + before/after verification"),
    ("WS7", "Pull historical CEG project sites for satellite validation case study", "CEG", "2026-06-15", "Not Started", "Independent third-party validation asset"),
    ("WS7", "Brenelli: legal framing for polluter outreach via satellite intel", "Lapo", "2026-06-30", "Not Started", "Avoid adversarial trigger"),

    ("WS8", "Request BioBlend + ISCO technical specs from CEG", "Lapo", "2026-06-15", "Not Started", "Subject to IP review"),
    ("WS8", "CEG: IP/licensing constraints assessment for third-party formulation", "CEG", "2026-06-30", "Not Started", "Has CEG done this in other markets?"),
    ("WS8", "RDC quality + handling capability assessment (ISO 9001/14001/45001)", "Lapo", "2026-07-31", "Not Started", "Verify they can hit CEG specs"),
    ("WS8", "Draft local manufacturing term sheet w/ RDC", "Lapo", "2026-08-15", "Not Started", "License / transfer / margin model"),

    ("WS9", "Hold Istanbul contact warm (no pitch)", "Lapo", "Ongoing", "In Progress", "DO NOT initiate before WS1 + WS6 closed"),
    ("WS9", "Draft Istanbul one-pager (regional partner framing)", "Lapo", "2026-07-15", "Not Started", "After commercial terms agreed"),
    ("WS9", "Initial Istanbul call (entity authority)", "Lapo", "2026-09-30", "Not Started", "Position as regional tech partner, not CEG-branded"),
]

# ----- Build workbook -----
wb = Workbook()

# ===== Sheet 1: Dashboard =====
ws = wb.active
ws.title = "Dashboard"

# Title
ws["A1"] = "CEG Italia — Project Management Dashboard"
ws["A1"].font = Font(name=FONT_NAME, bold=True, size=18, color=COL_HEADER_BG)
ws.merge_cells("A1:H1")

ws["A2"] = f"Last updated: {TODAY.strftime('%B %d, %Y')}  |  Owner: Lapo Lazzati  |  Counterpart: Stephen Tanico (CEG)"
ws["A2"].font = Font(name=FONT_NAME, italic=True, size=10, color="595959")
ws.merge_cells("A2:H2")

# KPI strip
kpi_row = 4
kpis = [
    ("Active Workstreams", f"=COUNTA(Workstreams!B4:B{4+len(WORKSTREAMS)-1})"),
    ("Open Action Items", f'=COUNTIF(\'Action Items\'!E4:E{4+len(ACTIONS)-1},"<>Done")'),
    ("Due This Week", f'=SUMPRODUCT((\'Action Items\'!D4:D{4+len(ACTIONS)-1}>=TODAY())*(\'Action Items\'!D4:D{4+len(ACTIONS)-1}<=TODAY()+7)*(\'Action Items\'!E4:E{4+len(ACTIONS)-1}<>"Done"))'),
    ("Overdue", f'=SUMPRODUCT((\'Action Items\'!D4:D{4+len(ACTIONS)-1}<TODAY())*(\'Action Items\'!E4:E{4+len(ACTIONS)-1}<>"Done")*(\'Action Items\'!D4:D{4+len(ACTIONS)-1}>0))'),
]
for i, (label, formula) in enumerate(kpis):
    col = 1 + i * 2
    cell_label = ws.cell(row=kpi_row, column=col, value=label)
    cell_value = ws.cell(row=kpi_row, column=col + 1, value=formula)
    cell_label.font = Font(name=FONT_NAME, bold=True, size=10, color="595959")
    cell_value.font = Font(name=FONT_NAME, bold=True, size=14, color=COL_HEADER_BG)
    cell_value.alignment = Alignment(horizontal="left")

# Workstream summary table
hdr_row = 6
headers = ["ID", "Workstream", "Priority", "Owner", "Status", "Deadline", "Start", "End"]
for i, h in enumerate(headers):
    c = ws.cell(row=hdr_row, column=i + 1, value=h)
    c.font = Font(name=FONT_NAME, bold=True, color=COL_HEADER_FG)
    c.fill = PatternFill("solid", start_color=COL_HEADER_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = BORDER_ALL

for i, w in enumerate(WORKSTREAMS):
    r = hdr_row + 1 + i
    start_d = WEEKS[w["start_w"]]
    end_d = WEEKS[min(w["end_w"], len(WEEKS) - 1)] + timedelta(days=4)
    row_vals = [
        w["id"], w["name"], w["priority"], w["owner"], w["status"],
        w["deadline"], start_d, end_d,
    ]
    for j, v in enumerate(row_vals):
        c = ws.cell(row=r, column=j + 1, value=v)
        c.font = Font(name=FONT_NAME, size=10)
        c.border = BORDER_ALL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if j in (6, 7):
            c.number_format = "yyyy-mm-dd"

# Column widths
widths_dash = {"A": 7, "B": 42, "C": 9, "D": 16, "E": 18, "F": 14, "G": 12, "H": 12}
for col, w in widths_dash.items():
    ws.column_dimensions[col].width = w

# Status legend
legend_row = hdr_row + len(WORKSTREAMS) + 3
ws.cell(row=legend_row, column=1, value="Status legend").font = Font(name=FONT_NAME, bold=True, size=11)
legend_items = [
    ("Done", COL_DONE_BG, "Completed"),
    ("In Progress", COL_INPROG_BG, "Active work"),
    ("Not Started", "FFFFFF", "Queued"),
    ("On Hold", "D9D9D9", "Paused / dependency"),
    ("Blocked", COL_BLOCKED_BG, "Blocked / overdue"),
]
for i, (label, color, desc) in enumerate(legend_items):
    r = legend_row + 1 + i
    cl = ws.cell(row=r, column=1, value=label)
    cl.fill = PatternFill("solid", start_color=color)
    cl.font = Font(name=FONT_NAME, bold=True, size=10)
    cl.border = BORDER_ALL
    cd = ws.cell(row=r, column=2, value=desc)
    cd.font = Font(name=FONT_NAME, size=10)

# Priority legend
ws.cell(row=legend_row, column=4, value="Priority legend").font = Font(name=FONT_NAME, bold=True, size=11)
prio_items = [
    ("P1", COL_BAR_PRIORITY, "Critical / blocking"),
    ("P2", COL_BAR_HIGH, "High - active sales"),
    ("P3", COL_BAR_MEDIUM, "Medium - infra"),
    ("P4", COL_BAR_NORMAL, "Normal - partnerships"),
    ("P5", COL_BAR_FUTURE, "Future / dependent"),
]
for i, (label, color, desc) in enumerate(prio_items):
    r = legend_row + 1 + i
    cl = ws.cell(row=r, column=4, value=label)
    cl.fill = PatternFill("solid", start_color=color)
    cl.font = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
    cl.border = BORDER_ALL
    cl.alignment = Alignment(horizontal="center")
    cd = ws.cell(row=r, column=5, value=desc)
    cd.font = Font(name=FONT_NAME, size=10)

ws.row_dimensions[1].height = 28
ws.row_dimensions[hdr_row].height = 22
ws.freeze_panes = "A7"

# ===== Sheet 2: Gantt =====
gws = wb.create_sheet("Gantt")

gws["A1"] = "CEG Italia — Gantt Chart"
gws["A1"].font = Font(name=FONT_NAME, bold=True, size=16, color=COL_HEADER_BG)
gws.merge_cells(f"A1:{get_column_letter(8 + N_WEEKS)}1")

gws["A2"] = f"Timeline: {WEEKS[0].strftime('%b %d, %Y')} — {(WEEKS[-1] + timedelta(days=4)).strftime('%b %d, %Y')}  ·  Each column = 1 week"
gws["A2"].font = Font(name=FONT_NAME, italic=True, size=10, color="595959")
gws.merge_cells(f"A2:{get_column_letter(8 + N_WEEKS)}2")

# Header rows
hdr1 = 4  # month band
hdr2 = 5  # week start dates
hdr_left = ["ID", "Workstream", "Priority", "Owner", "Status", "Start", "End"]
for i, h in enumerate(hdr_left):
    for r in (hdr1, hdr2):
        c = gws.cell(row=r, column=i + 1, value=h if r == hdr2 else None)
        c.fill = PatternFill("solid", start_color=COL_HEADER_BG)
        c.border = BORDER_ALL
    cell = gws.cell(row=hdr2, column=i + 1)
    cell.font = Font(name=FONT_NAME, bold=True, color=COL_HEADER_FG, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")

# Merge month bands
prev_month = None
month_start_col = 8
for i, wk in enumerate(WEEKS):
    col = 8 + i
    month_str = wk.strftime("%b %Y")
    cell_top = gws.cell(row=hdr1, column=col, value=month_str if month_str != prev_month else None)
    cell_top.fill = PatternFill("solid", start_color=COL_HEADER_BG)
    cell_top.font = Font(name=FONT_NAME, bold=True, color=COL_HEADER_FG, size=10)
    cell_top.alignment = Alignment(horizontal="center")
    cell_top.border = BORDER_ALL

    cell_btm = gws.cell(row=hdr2, column=col, value=wk)
    cell_btm.number_format = "mm-dd"
    cell_btm.fill = PatternFill("solid", start_color=COL_SUBHEADER_BG)
    cell_btm.font = Font(name=FONT_NAME, bold=True, size=9, color="1F4E78")
    cell_btm.alignment = Alignment(horizontal="center")
    cell_btm.border = BORDER_ALL
    prev_month = month_str

# Workstream rows
priority_color = {
    "P1": COL_BAR_PRIORITY, "P2": COL_BAR_HIGH, "P3": COL_BAR_MEDIUM,
    "P4": COL_BAR_NORMAL, "P5": COL_BAR_FUTURE, "P6": COL_BAR_FUTURE,
}

start_data = hdr2 + 1
for i, w in enumerate(WORKSTREAMS):
    r = start_data + i
    sd = WEEKS[w["start_w"]]
    ed = WEEKS[min(w["end_w"], len(WEEKS) - 1)] + timedelta(days=4)
    left = [w["id"], w["name"], w["priority"], w["owner"], w["status"], sd, ed]
    for j, v in enumerate(left):
        c = gws.cell(row=r, column=j + 1, value=v)
        c.font = Font(name=FONT_NAME, size=10)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BORDER_ALL
        if j in (5, 6):
            c.number_format = "yyyy-mm-dd"

    # Bar
    bar_color = priority_color.get(w["priority"], COL_BAR_NORMAL)
    for k in range(w["start_w"], min(w["end_w"] + 1, N_WEEKS)):
        col = 8 + k
        cell = gws.cell(row=r, column=col)
        cell.fill = PatternFill("solid", start_color=bar_color)
        cell.border = BORDER_ALL
        # Label start of bar with workstream short tag
        if k == w["start_w"]:
            cell.value = w["priority"]
            cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=8)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Empty cells - thin border
    for k in range(N_WEEKS):
        col = 8 + k
        cell = gws.cell(row=r, column=col)
        if cell.fill.fgColor.rgb in (None, "00000000"):
            cell.border = BORDER_ALL

# Today marker (red column highlight)
today_week_idx = (TODAY - WEEK_START).days // 7
today_col = 8 + today_week_idx
for r in range(hdr1, start_data + len(WORKSTREAMS)):
    cell = gws.cell(row=r, column=today_col)
    existing_fill = cell.fill.fgColor.rgb if cell.fill.fgColor else None
    # Add red top/bottom border to mark today's column
    cell.border = Border(
        left=Side(border_style="medium", color=COL_TODAY_LINE),
        right=Side(border_style="medium", color=COL_TODAY_LINE),
        top=cell.border.top,
        bottom=cell.border.bottom,
    )

# Today label above
gws.cell(row=3, column=today_col, value="◀ TODAY").font = Font(name=FONT_NAME, bold=True, color=COL_TODAY_LINE, size=9)
gws.cell(row=3, column=today_col).alignment = Alignment(horizontal="center")

# Column widths
gws.column_dimensions["A"].width = 6
gws.column_dimensions["B"].width = 38
gws.column_dimensions["C"].width = 8
gws.column_dimensions["D"].width = 14
gws.column_dimensions["E"].width = 16
gws.column_dimensions["F"].width = 11
gws.column_dimensions["G"].width = 11
for k in range(N_WEEKS):
    gws.column_dimensions[get_column_letter(8 + k)].width = 5

gws.row_dimensions[hdr1].height = 18
gws.row_dimensions[hdr2].height = 22
for i in range(len(WORKSTREAMS)):
    gws.row_dimensions[start_data + i].height = 28

gws.freeze_panes = "H6"

# ===== Sheet 3: Workstreams =====
wws = wb.create_sheet("Workstreams")

wws["A1"] = "Workstream Definitions"
wws["A1"].font = Font(name=FONT_NAME, bold=True, size=16, color=COL_HEADER_BG)
wws.merge_cells("A1:H1")

ws_hdr = ["ID", "Workstream", "Priority", "Owner", "Status", "Deadline", "Start", "End", "Notes / Strategic Context"]
for i, h in enumerate(ws_hdr):
    c = wws.cell(row=3, column=i + 1, value=h)
    c.font = Font(name=FONT_NAME, bold=True, color=COL_HEADER_FG)
    c.fill = PatternFill("solid", start_color=COL_HEADER_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = BORDER_ALL

for i, w in enumerate(WORKSTREAMS):
    r = 4 + i
    sd = WEEKS[w["start_w"]]
    ed = WEEKS[min(w["end_w"], len(WEEKS) - 1)] + timedelta(days=4)
    vals = [w["id"], w["name"], w["priority"], w["owner"], w["status"], w["deadline"], sd, ed, w["notes"]]
    for j, v in enumerate(vals):
        c = wws.cell(row=r, column=j + 1, value=v)
        c.font = Font(name=FONT_NAME, size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = BORDER_ALL
        if j in (6, 7):
            c.number_format = "yyyy-mm-dd"

widths_ws = {"A": 7, "B": 36, "C": 9, "D": 16, "E": 18, "F": 14, "G": 12, "H": 12, "I": 60}
for col, w in widths_ws.items():
    wws.column_dimensions[col].width = w
for i in range(len(WORKSTREAMS)):
    wws.row_dimensions[4 + i].height = 50

wws.freeze_panes = "A4"

# ===== Sheet 4: Action Items =====
aws = wb.create_sheet("Action Items")

aws["A1"] = "Action Items — Granular Tasks"
aws["A1"].font = Font(name=FONT_NAME, bold=True, size=16, color=COL_HEADER_BG)
aws.merge_cells("A1:G1")

ai_hdr = ["WS", "Task", "Owner", "Due Date", "Status", "Days to Due", "Notes"]
for i, h in enumerate(ai_hdr):
    c = aws.cell(row=3, column=i + 1, value=h)
    c.font = Font(name=FONT_NAME, bold=True, color=COL_HEADER_FG)
    c.fill = PatternFill("solid", start_color=COL_HEADER_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = BORDER_ALL

for i, (ws_id, task, owner, due, status, notes) in enumerate(ACTIONS):
    r = 4 + i
    # Parse due date
    due_val = due
    is_date = False
    try:
        due_val = date.fromisoformat(due)
        is_date = True
    except Exception:
        pass

    vals = [ws_id, task, owner, due_val, status]
    for j, v in enumerate(vals):
        c = aws.cell(row=r, column=j + 1, value=v)
        c.font = Font(name=FONT_NAME, size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = BORDER_ALL
        if j == 3 and is_date:
            c.number_format = "yyyy-mm-dd"

    # Days to due (formula)
    if is_date:
        c_days = aws.cell(row=r, column=6, value=f"=D{r}-TODAY()")
    else:
        c_days = aws.cell(row=r, column=6, value="—")
    c_days.font = Font(name=FONT_NAME, size=10)
    c_days.alignment = Alignment(horizontal="center", vertical="top")
    c_days.border = BORDER_ALL
    c_days.number_format = "0"

    c_notes = aws.cell(row=r, column=7, value=notes)
    c_notes.font = Font(name=FONT_NAME, size=9, italic=True, color="595959")
    c_notes.alignment = Alignment(vertical="top", wrap_text=True)
    c_notes.border = BORDER_ALL

# Conditional formatting on Status column (E)
status_range = f"E4:E{4 + len(ACTIONS) - 1}"
aws.conditional_formatting.add(
    status_range,
    FormulaRule(formula=[f'$E4="Done"'], fill=PatternFill("solid", start_color=COL_DONE_BG)),
)
aws.conditional_formatting.add(
    status_range,
    FormulaRule(formula=[f'$E4="In Progress"'], fill=PatternFill("solid", start_color=COL_INPROG_BG)),
)
aws.conditional_formatting.add(
    status_range,
    FormulaRule(formula=[f'$E4="Blocked"'], fill=PatternFill("solid", start_color=COL_BLOCKED_BG)),
)
aws.conditional_formatting.add(
    status_range,
    FormulaRule(formula=[f'$E4="On Hold"'], fill=PatternFill("solid", start_color="D9D9D9")),
)

# Conditional formatting on Days to Due column (F): red if <0 and not Done
days_range = f"F4:F{4 + len(ACTIONS) - 1}"
aws.conditional_formatting.add(
    days_range,
    FormulaRule(formula=[f'AND(ISNUMBER($F4),$F4<0,$E4<>"Done")'],
                fill=PatternFill("solid", start_color=COL_BLOCKED_BG),
                font=Font(name=FONT_NAME, size=10, bold=True, color="9C0006")),
)
aws.conditional_formatting.add(
    days_range,
    FormulaRule(formula=[f'AND(ISNUMBER($F4),$F4>=0,$F4<=7,$E4<>"Done")'],
                fill=PatternFill("solid", start_color=COL_INPROG_BG)),
)

widths_ai = {"A": 6, "B": 60, "C": 16, "D": 12, "E": 16, "F": 10, "G": 50}
for col, w in widths_ai.items():
    aws.column_dimensions[col].width = w
for i in range(len(ACTIONS)):
    aws.row_dimensions[4 + i].height = 38

aws.freeze_panes = "A4"
aws.auto_filter.ref = f"A3:G{3 + len(ACTIONS)}"

# Save
out_path = "/sessions/relaxed-amazing-einstein/mnt/outputs/CEG_Italia_Project_Dashboard.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Workstreams: {len(WORKSTREAMS)}")
print(f"Action items: {len(ACTIONS)}")
print(f"Timeline: {WEEKS[0]} to {WEEKS[-1] + timedelta(days=4)}")
